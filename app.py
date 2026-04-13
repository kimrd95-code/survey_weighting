"""
Веб-сервис взвешивания опросов (Streamlit).
Режим 1 — подгонка к Росстату: вес по ячейке = целевая доля / фактическая (без нормировки к среднему 1).
Режим 2 — группа на группу: внутренняя калибровка; у эталона (целевой группы) вес фиксированно 1, остальные масштабируются.
"""

from __future__ import annotations

import copy
import io
import re
import hashlib
from typing import Any, Optional

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

# --- Константы ---
TARGET_COL_CANDIDATES = (
    "target",
    "доля",
    "share",
    "proportion",
    "prop",
    "p",
    "count",
    "численность",
    "население",
    "n",
    "freq",
)
RAW_LOW = 0.3
RAW_HIGH = 2.5
EXTREME_BG = "background-color: #ffcccc"

# Режим 1, табличный Росстат: фиксированные колонки опроса (первая строка — имена переменных)
MODE1_STANDARD_SURVEY_QGENDER = "Qgender"
MODE1_STANDARD_SURVEY_QAGERANGE = "Qagerange"
MODE1_STANDARD_SURVEY_QCITYSIZE = "Qrucitysize"
# Сопоставление измерений Росстата → колонка опроса (имена в файле Росстата — типичные варианты)
MODE1_STANDARD_TABLE_SPEC: list[tuple[tuple[str, ...], str]] = [
    (("Пол", "пол"), MODE1_STANDARD_SURVEY_QGENDER),
    (("Возраст", "возраст"), MODE1_STANDARD_SURVEY_QAGERANGE),
    (
        ("Тип населённого пункта", "Тип НП", "тип населенного пункта"),
        MODE1_STANDARD_SURVEY_QCITYSIZE,
    ),
]


def _init_session() -> None:
    defaults = {
        "survey_df": None,
        "survey_name": None,
        "survey_header_row": 1,
        "survey_bytes": None,
        "rosstat_df": None,
        "rosstat_name": None,
        "weights": None,
        "preview_df": None,
        "merge_edges": [],
        "mode1_recalc": None,
        "mode2_params": None,
        "preview_mode": None,
        "active_mode": None,
        "preview_baseline": None,
        "last_info": [],
        "last_warnings": [],
        "survey_leading_empty_rows": 0,
        "mode1_dim_merge_groups": {},
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def reset_session() -> None:
    for k in list(st.session_state.keys()):
        del st.session_state[k]
    _init_session()


def normalize_key(x: Any) -> str:
    if pd.isna(x):
        return ""
    if isinstance(x, (int, np.integer)):
        return str(int(x))
    if isinstance(x, (float, np.floating)):
        if np.isnan(float(x)):
            return ""
        xf = float(x)
        if xf.is_integer():
            return str(int(xf))
        t = str(xf).strip().lower()
    else:
        t = str(x).strip().lower()
    t = t.replace("ё", "е")
    for a, b in [("–", "-"), ("—", "-"), ("−", "-")]:
        t = t.replace(a, b)
    t = " ".join(t.split())
    return t


def is_no_text_cell(x: Any) -> bool:
    """True, если в ячейке нет ответа: пропуск, пустая строка или только пробелы."""
    if pd.isna(x):
        return True
    if isinstance(x, str):
        return x.strip() == ""
    try:
        s = str(x).strip()
    except Exception:
        return True
    return s == ""


def row_is_fully_empty(row: pd.Series) -> bool:
    """True, если во всех колонках строки нет данных (как полностью пустая строка в Excel)."""
    return bool(all(is_no_text_cell(v) for v in row))


def count_leading_fully_empty_rows(df: pd.DataFrame) -> int:
    """Сколько полностью пустых строк подряд с начала таблицы (до первой непустой)."""
    n = 0
    for i in range(len(df)):
        if row_is_fully_empty(df.iloc[i]):
            n += 1
        else:
            break
    return n


def drop_fully_empty_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Удаляет строки, где все ячейки пустые — их не считаем респондентами."""
    if df.empty:
        return df
    keep = ~df.apply(row_is_fully_empty, axis=1)
    return df.loc[keep].reset_index(drop=True)


def survey_cell_stratum_label(x: Any) -> str:
    """Метка для склейки страт: без текста → __ПРОПУСК__, иначе текст без лишних пробелов по краям."""
    if is_no_text_cell(x):
        return "__ПРОПУСК__"
    if isinstance(x, str):
        return x.strip()
    return str(x).strip()


def category_label_for_ui(x: Any) -> str:
    """Читаемая подпись категории для списков (Excel часто даёт 1.0 вместо 1)."""
    if is_no_text_cell(x):
        return ""
    if isinstance(x, (float, np.floating)):
        if not np.isfinite(float(x)):
            return ""
        xf = float(x)
        if xf.is_integer():
            return str(int(xf))
        return str(x).strip()
    if isinstance(x, (int, np.integer)):
        return str(int(x))
    if isinstance(x, str):
        return x.strip()
    return str(x).strip()


def _find_rosstat_dim_column(df: pd.DataFrame, aliases: tuple[str, ...]) -> Optional[str]:
    for c in df.columns:
        sc = str(c).strip()
        for a in aliases:
            if sc == a or normalize_key(sc) == normalize_key(a):
                return c
    return None


def _survey_column_for_standard_var(survey: pd.DataFrame, std_name: str) -> Optional[str]:
    """Колонка в df по метке режима 1 (Qgender и т.д.) в строке файла над заголовком переменных."""
    want = std_name.strip()
    row = survey.attrs.get("survey_var_row") or []
    for j, cell in enumerate(row):
        if j >= len(survey.columns):
            break
        if is_no_text_cell(cell):
            continue
        lab = str(cell).strip()
        if lab == want or normalize_key(lab) == normalize_key(want):
            return str(survey.columns[j])
    if want in survey.columns:
        return want
    return None


def resolve_mode1_standard_table(
    ros: pd.DataFrame, survey: pd.DataFrame
) -> tuple[Optional[list[str]], Optional[dict[str, str]], Optional[str]]:
    """Три измерения Росстата (как сейчас) и Qgender / Qagerange / Qrucitysize в первой строке опроса над заголовком."""
    ros_dims: list[str] = []
    survey_map: dict[str, str] = {}
    missing_ros: list[str] = []
    missing_sur: list[str] = []
    for aliases, std_var in MODE1_STANDARD_TABLE_SPEC:
        rcol = _find_rosstat_dim_column(ros, aliases)
        if rcol is None:
            missing_ros.append(aliases[0])
            continue
        scol = _survey_column_for_standard_var(survey, std_var)
        if scol is None:
            missing_sur.append(std_var)
            continue
        ros_dims.append(rcol)
        survey_map[rcol] = scol
    if missing_ros or missing_sur or len(ros_dims) != 3:
        parts: list[str] = []
        if missing_ros:
            parts.append("в файле Росстата не найдены колонки: " + ", ".join(missing_ros))
        if missing_sur:
            parts.append(
                "в первой строке опроса (над строкой с названиями колонок) нет меток: "
                + ", ".join(missing_sur)
                + " (либо совпадения с именем колонки во второй строке)"
            )
        return None, None, "Стандартное сопоставление недоступно. " + "; ".join(parts) + "."
    return ros_dims, survey_map, None


def _unique_ui_categories(series: pd.Series) -> list[str]:
    u: set[str] = set()
    for x in series.dropna():
        lab = category_label_for_ui(x)
        if lab:
            u.add(lab)
    return sorted(u, key=natural_sort_key)


def _norm_keys_for_merged_groups(dim: str, groups: list[list[str]] | None) -> set[str]:
    out: set[str] = set()
    for g in groups or []:
        for x in g:
            out.add(normalize_key(category_label_for_ui(x)))
            out.add(normalize_key(str(x).strip()))
    return out


def stratum_is_proskip_only(s: str) -> bool:
    """Страта только из пустых соц-дем (__ПРОПУСК__ / варианты с подчёркиваниями); в предпросмотр не выводим, вес не калибруем."""
    t = str(s).strip()
    if not t:
        return False
    for piece in t.split(","):
        inner = piece.strip().strip("_").lower().replace("ё", "е")
        if inner != "пропуск":
            return False
    return True


def load_survey_excel(uploaded: Any, header_row: int) -> pd.DataFrame:
    uploaded.seek(0)
    hr = int(header_row)
    df = pd.read_excel(uploaded, sheet_name=0, header=hr)
    survey_var_row: list[str] = []
    if hr >= 1:
        uploaded.seek(0)
        raw_top = pd.read_excel(uploaded, sheet_name=0, header=None, nrows=hr)
        ncols = len(df.columns)
        if len(raw_top) > hr - 1:
            arr = raw_top.iloc[hr - 1].values
            for j in range(ncols):
                if j < len(arr):
                    v = arr[j]
                    survey_var_row.append("" if pd.isna(v) else str(v).strip())
                else:
                    survey_var_row.append("")
        else:
            survey_var_row = [""] * ncols
    else:
        survey_var_row = [""] * len(df.columns)
    leading = count_leading_fully_empty_rows(df)
    df = drop_fully_empty_rows(df)
    df.attrs["leading_fully_empty_rows"] = leading
    df.attrs["survey_var_row"] = survey_var_row
    return df


def read_rosstat_excel(uploaded: Any) -> pd.DataFrame:
    uploaded.seek(0)
    raw = pd.read_excel(uploaded, sheet_name=0, header=None)
    if raw.shape[0] >= 2:
        row0 = str(raw.iloc[0, 0]).strip().lower() if pd.notna(raw.iloc[0, 0]) else ""
        row1 = str(raw.iloc[1, 0]).strip().lower() if pd.notna(raw.iloc[1, 0]) else ""
        if "страта" in row0 or (row0 == "" and row1 and not row1.replace(".", "").replace("-", "").isdigit()):
            uploaded.seek(0)
            return pd.read_excel(uploaded, sheet_name=0, header=1)
    uploaded.seek(0)
    return pd.read_excel(uploaded, sheet_name=0, header=0)


def detect_rosstat_matrix(df: pd.DataFrame) -> bool:
    if df.shape[1] < 2:
        return False
    first = df.iloc[:, 0]
    rest = df.iloc[:, 1:]
    if not rest.apply(pd.to_numeric, errors="coerce").notna().all().all():
        return False
    return first.dtype == object or str(first.dtype).startswith("str")


def melt_matrix_rosstat(df: pd.DataFrame) -> pd.DataFrame:
    row_name = str(df.columns[0])
    rows = []
    for _, r in df.iterrows():
        k = r.iloc[0]
        if pd.isna(k):
            continue
        for c in df.columns[1:]:
            v = r[c]
            if pd.notna(v) and float(v) != 0:
                rows.append({row_name: k, "dim_col": str(c), "count": float(v)})
    return pd.DataFrame(rows)


def _pick_target_col(df: pd.DataFrame) -> Optional[str]:
    lower = {str(c).lower(): c for c in df.columns}
    for c in TARGET_COL_CANDIDATES:
        if c in lower:
            return lower[c]
    num = df.select_dtypes(include=[np.number]).columns
    if len(num) == 1:
        return num[0]
    if len(num) > 1:
        return num[-1]
    return None


NO_MERGE_STRATUM = "— Не объединять —"


def merge_edges_list_to_set(edges: list | None) -> set[frozenset[str]]:
    out: set[frozenset[str]] = set()
    for e in edges or []:
        if len(e) >= 2:
            out.add(frozenset({str(e[0]), str(e[1])}))
    return out


class _UnionFindStrata:
    def __init__(self, nodes: set[str]):
        self.parent = {x: x for x in nodes}

    def find(self, x: str) -> str:
        if self.parent[x] != x:
            self.parent[x] = self.find(self.parent[x])
        return self.parent[x]

    def union(self, a: str, b: str) -> None:
        ra, rb = self.find(a), self.find(b)
        if ra != rb:
            self.parent[rb] = ra


class _UnionFindInt:
    def __init__(self, n: int):
        self.parent = list(range(n))

    def find(self, x: int) -> int:
        if self.parent[x] != x:
            self.parent[x] = self.find(self.parent[x])
        return self.parent[x]

    def union(self, a: int, b: int) -> None:
        ra, rb = self.find(a), self.find(b)
        if ra != rb:
            self.parent[rb] = ra


def merge_groups_to_edges(groups: list[list[str]] | None) -> list[list[str]]:
    """Список групп категорий → рёбра для union-find (цепочка внутри группы)."""
    edges: list[list[str]] = []
    for g in groups or []:
        g2 = [str(x).strip() for x in g if str(x).strip()]
        if len(g2) < 2:
            continue
        for i in range(len(g2) - 1):
            edges.append([g2[i], g2[i + 1]])
    return edges


def normalized_dimension_merge_map(universe_norm: set[str], merge_edges: list | None) -> dict[str, str]:
    """Ключи уже в пространстве normalize_key; возвращает norm → укрупнённая метка (тоже в нижнем регистре и т.д.)."""
    out: dict[str, str] = {k: k for k in universe_norm}
    edge_set = merge_edges_list_to_set(merge_edges)
    if not edge_set:
        return out
    nodes: set[str] = set(universe_norm)
    for p in edge_set:
        for x in p:
            nodes.add(normalize_key(str(x)))
    uf = _UnionFindStrata(nodes)
    for p in edge_set:
        a, b = tuple(p)
        uf.union(normalize_key(str(a)), normalize_key(str(b)))
    root_members: dict[str, list[str]] = {}
    for n in nodes:
        r = uf.find(n)
        root_members.setdefault(r, []).append(n)
    label_for: dict[str, str] = {}
    for members in root_members.values():
        if len(members) == 1:
            m0 = members[0]
            label_for[m0] = m0
        else:
            lbl = " + ".join(sorted(members, key=natural_sort_key))
            for m in members:
                label_for[m] = lbl
    for k in universe_norm:
        out[k] = label_for.get(k, k)
    return out


def apply_merge_edges(sk: pd.Series, merge_edges: list | None) -> pd.Series:
    """Объединяет страты по рёбрам; имя группы — «A + B + …» (порядок по natural_sort_key)."""
    edge_set = merge_edges_list_to_set(merge_edges)
    if not edge_set:
        return sk.astype(str).copy()
    sks = sk.astype(str)
    nodes: set[str] = set(sks.unique())
    for p in edge_set:
        nodes |= set(p)
    uf = _UnionFindStrata(nodes)
    for p in edge_set:
        a, b = tuple(p)
        uf.union(a, b)
    root_members: dict[str, list[str]] = {}
    for n in nodes:
        r = uf.find(n)
        root_members.setdefault(r, []).append(n)
    label_for: dict[str, str] = {}
    for members in root_members.values():
        if len(members) == 1:
            m0 = members[0]
            label_for[m0] = m0
        else:
            lbl = " + ".join(sorted(members, key=natural_sort_key))
            for m in members:
                label_for[m] = lbl
    return sks.map(lambda v: label_for.get(v, v))


def build_stratum_key(df: pd.DataFrame, cols: list[str]) -> pd.Series:
    parts = []
    for c in cols:
        parts.append(df[c].map(survey_cell_stratum_label))
    if len(parts) == 1:
        return parts[0]
    return parts[0].str.cat(parts[1:], sep=", ")


def matrix_with_exclusions(
    rosstat: pd.DataFrame,
    exclude_rows: list[str],
    exclude_cols: list[str],
) -> tuple[pd.DataFrame, float, str]:
    joint = melt_matrix_rosstat(rosstat)
    if joint.empty:
        raise ValueError("В матрице Росстата нет числовых данных для расчёта.")
    rname = str(rosstat.columns[0])
    joint["_kr"] = joint[rname].map(normalize_key)
    joint["_kc"] = joint["dim_col"].map(normalize_key)
    exr = {normalize_key(x) for x in exclude_rows}
    exc = {normalize_key(x) for x in exclude_cols}
    if exr or exc:
        joint = joint.loc[(~joint["_kr"].isin(exr)) & (~joint["_kc"].isin(exc))].copy()
    total = float(joint["count"].sum())
    if total <= 0:
        raise ValueError("После исключения групп сумма целевых значений равна нулю.")
    msg = ""
    if exr or exc:
        msg = "Часть групп Росстата исключена из расчёта целевых долей."
    return joint, total, msg


def check_and_normalize_target_shares(pt_map: dict[tuple[str, str], float], infos: list[str]) -> dict[tuple[str, str], float]:
    s = sum(pt_map.values())
    if s <= 0:
        return pt_map
    if abs(s - 1.0) > 1e-6:
        infos.append("Целевые доли Росстата не суммируются в 1, выполнена автоматическая нормировка.")
        return {k: float(v) / s for k, v in pt_map.items()}
    return pt_map


def _format_mode1_cell_key(k: tuple[str, ...]) -> str:
    if len(k) == 2:
        return f"«{k[0]} / {k[1]}»"
    return "«" + " | ".join(str(x) for x in k) + "»"


def _mode1_preview_label_for_group(
    idxs: list[int],
    stratum_per_row: list[str],
    row_cell_keys: list[tuple[str, ...]],
) -> str:
    """Подпись строки предпросмотра: укрупнённая страта опроса или ключ ячейки Росстата."""
    slabs = {str(stratum_per_row[i]) for i in idxs}
    if len(slabs) == 1:
        return next(iter(slabs))
    k0 = row_cell_keys[idxs[0]]
    if all(row_cell_keys[i] == k0 for i in idxs):
        return " | ".join(str(x) for x in k0)
    return " + ".join(sorted(slabs, key=natural_sort_key))


def direct_weights_mode1_poststrat(
    n: int,
    stratum_per_row: list[str],
    row_cell_keys: list[tuple[str, ...]],
    pt_map: dict[tuple[str, ...], float],
) -> tuple[np.ndarray, list[dict[str, Any]], list[str], list[str]]:
    """Веса и предпросмотр по группам респондентов, объединённым так:

    - все с **одной целевой ячейкой** Росстата (после укрупнения измерений) — общий вес;
    - все с **одной укрупнённой стратой опроса** (после «Объединить страты») — одна строка предпросмотра, цели суммируются по уникальным ячейкам.

    Компоненты связности: ребро между каждой парой строк с одинаковым ``stratum_per_row`` и между парами с одинаковым ``row_cell_key``.
    """
    warns: list[str] = []
    infos: list[str] = []
    w = np.ones(n, dtype=float)
    preview_rows: list[dict[str, Any]] = []

    idx_proskip: dict[str, list[int]] = {}
    active: list[int] = []
    for i in range(n):
        s = str(stratum_per_row[i])
        if stratum_is_proskip_only(s):
            idx_proskip.setdefault(s, []).append(i)
        else:
            active.append(i)

    for s, idxs in sorted(idx_proskip.items(), key=lambda x: str(x[0])):
        for i in idxs:
            w[i] = 1.0
        n_s = len(idxs)
        p_a = n_s / n if n else 0.0
        preview_rows.append(
            {
                "stratum": s,
                "n": n_s,
                "share_act": p_a,
                "share_tgt": 0.0,
                "raw_w": 1.0,
            }
        )

    if not active:
        return w, preview_rows, warns, infos

    uf = _UnionFindInt(n)
    by_s: dict[str, list[int]] = {}
    by_k: dict[tuple[str, ...], list[int]] = {}
    for i in active:
        by_s.setdefault(str(stratum_per_row[i]), []).append(i)
        by_k.setdefault(row_cell_keys[i], []).append(i)
    for lst in by_s.values():
        for j in range(1, len(lst)):
            uf.union(lst[0], lst[j])
    for lst in by_k.values():
        for j in range(1, len(lst)):
            uf.union(lst[0], lst[j])

    components: dict[int, list[int]] = {}
    for i in active:
        r = uf.find(i)
        components.setdefault(r, []).append(i)

    for idxs in sorted(components.values(), key=lambda ix: _mode1_preview_label_for_group(ix, stratum_per_row, row_cell_keys)):
        keys_here: list[tuple[str, ...]] = []
        seen_k: set[tuple[str, ...]] = set()
        for i in idxs:
            k = row_cell_keys[i]
            if k not in seen_k:
                seen_k.add(k)
                keys_here.append(k)

        tgt_share = 0.0
        missing_any = False
        for k in keys_here:
            if k in pt_map:
                tgt_share += float(pt_map[k])
            else:
                missing_any = True
                if any(k):
                    warns.append(
                        "Не найдено соответствие между категориями в данных и целевом файле "
                        f"для сочетания {_format_mode1_cell_key(k)}."
                    )

        n_g = len(idxs)
        p_a = n_g / n if n else 0.0
        plab = _mode1_preview_label_for_group(idxs, stratum_per_row, row_cell_keys)

        if missing_any and tgt_share == 0 and n_g > 0:
            for i in idxs:
                w[i] = np.nan
            preview_rows.append(
                {
                    "stratum": plab,
                    "n": n_g,
                    "share_act": p_a,
                    "share_tgt": 0.0,
                    "raw_w": np.nan,
                }
            )
            continue

        if p_a <= 0:
            continue

        if tgt_share <= 0 and n_g > 0:
            warns.append(
                f"Для группы «{plab}» в целевом распределении суммарная доля равна нулю, но в выборке есть респонденты."
            )
            raw_w = np.nan
            for i in idxs:
                w[i] = np.nan
        else:
            raw_w = tgt_share / p_a if tgt_share > 0 else (0.0 if n_g == 0 else np.nan)
            for i in idxs:
                w[i] = raw_w

        preview_rows.append(
            {
                "stratum": plab,
                "n": n_g,
                "share_act": p_a,
                "share_tgt": tgt_share,
                "raw_w": float(raw_w) if np.isfinite(raw_w) else np.nan,
            }
        )

    if np.any(~np.isfinite(w)):
        return w, preview_rows, warns, infos

    return w, preview_rows, warns, infos


def mode1_matrix_compute(
    survey: pd.DataFrame,
    rosstat: pd.DataFrame,
    survey_row_cols: list[str],
    survey_col_col: str,
    merge_edges: list | None,
    exclude_rows: list[str],
    exclude_cols: list[str],
    dim_merge_groups: dict[str, list[list[str]]] | None = None,
) -> tuple[Optional[np.ndarray], Optional[pd.DataFrame], str, list[str], list[str]]:
    warns: list[str] = []
    infos: list[str] = []
    try:
        joint, _total_old, ex_msg = matrix_with_exclusions(rosstat, exclude_rows, exclude_cols)
    except ValueError as e:
        return None, None, str(e), [], []

    if ex_msg:
        infos.append(ex_msg)

    n = len(survey)
    if n == 0:
        return None, None, "Файл с опросом не содержит строк.", [], []

    d = survey.copy()
    d["_kr"] = build_stratum_key(d, survey_row_cols).map(normalize_key)
    d["_kc"] = d[survey_col_col].map(normalize_key)
    base_sk = build_stratum_key(d, survey_row_cols + [survey_col_col])
    d["_sk"] = apply_merge_edges(base_sk, merge_edges)

    dmg = dim_merge_groups or {}
    row_e = merge_groups_to_edges(dmg.get("matrix_row", []))
    col_e = merge_groups_to_edges(dmg.get("matrix_col", []))
    if row_e or col_e:
        univ_kr = set(joint["_kr"].astype(str)) | set(d["_kr"].astype(str))
        univ_kc = set(joint["_kc"].astype(str)) | set(d["_kc"].astype(str))
        map_kr = normalized_dimension_merge_map(univ_kr, row_e)
        map_kc = normalized_dimension_merge_map(univ_kc, col_e)
        joint = joint.copy()
        joint["_kr"] = joint["_kr"].astype(str).map(lambda x: map_kr.get(x, x))
        joint["_kc"] = joint["_kc"].astype(str).map(lambda x: map_kc.get(x, x))
        d["_kr"] = d["_kr"].astype(str).map(lambda x: map_kr.get(x, x))
        d["_kc"] = d["_kc"].astype(str).map(lambda x: map_kc.get(x, x))
        infos.append("Учтено укрупнение категорий Росстата по строкам и/или столбцам матрицы.")

    jt = joint.groupby(["_kr", "_kc"], dropna=False)["count"].sum()
    total_j = float(jt.sum())
    pt_map = {k: float(v) / float(total_j) for k, v in jt.items()}
    pt_map = check_and_normalize_target_shares(pt_map, infos)

    stratum_per_row = d["_sk"].astype(str).tolist()
    row_cell_keys = list(zip(d["_kr"].tolist(), d["_kc"].tolist()))

    w, preview_list, warns2, infos2 = direct_weights_mode1_poststrat(
        n, stratum_per_row, row_cell_keys, pt_map
    )
    warns.extend(warns2)
    infos.extend(infos2)

    if np.any(~np.isfinite(w)):
        return (
            None,
            None,
            "Для части респондентов нельзя рассчитать вес. Проверьте соответствие категорий в опросе и файле Росстата или объедините страты.",
            warns,
            infos,
        )

    preview = pd.DataFrame(preview_list)
    return w, preview, "", warns, infos


def mode1_table_compute(
    survey: pd.DataFrame,
    rosstat: pd.DataFrame,
    weight_vars: list[str],
    survey_col_map: dict[str, str],
    merge_edges: list | None,
    exclude_values: dict[str, list[str]],
    dim_merge_groups: dict[str, list[list[str]]] | None = None,
) -> tuple[Optional[np.ndarray], Optional[pd.DataFrame], str, list[str], list[str]]:
    warns: list[str] = []
    infos: list[str] = []
    if not weight_vars:
        return None, None, "Выберите переменные для взвешивания.", [], []

    n = len(survey)
    if n == 0:
        return None, None, "Пустая выборка.", [], []

    tgt_col = _pick_target_col(rosstat)
    if tgt_col is None:
        return None, None, "Не удалось определить столбец с целевыми значениями в файле Росстата.", [], []

    for v in weight_vars:
        if v not in rosstat.columns:
            return None, None, f"В файле Росстата нет колонки «{v}».", [], []

    d = survey.copy()
    rstat = rosstat.copy()
    for v in weight_vars:
        rstat[v] = rstat[v].map(normalize_key)

    m = np.ones(len(rstat), dtype=bool)
    for v, excl in exclude_values.items():
        if v in rstat.columns and excl:
            nv = {normalize_key(x) for x in excl}
            m &= ~rstat[v].isin(nv)
    rstat_f = rstat.loc[m].copy()
    if rstat_f.empty:
        return None, None, "После исключения групп в файле Росстата не осталось строк.", [], []

    norm_cols: list[str] = []
    for v in weight_vars:
        sc = survey_col_map[v]
        if sc not in d.columns:
            return None, None, f"В данных опроса нет колонки «{sc}».", [], []
        nc = f"_n_{v}"
        d[nc] = d[sc].map(normalize_key)
        norm_cols.append(nc)

    dmg = dim_merge_groups or {}
    dim_merged_any = False
    for v in weight_vars:
        edges = merge_groups_to_edges(dmg.get(v, []))
        if not edges:
            continue
        nc = f"_n_{v}"
        univ = set(rstat_f[v].astype(str).unique()) | set(d[nc].astype(str).unique())
        mmap = normalized_dimension_merge_map(univ, edges)
        rstat_f[v] = rstat_f[v].astype(str).map(lambda x: mmap.get(x, x))
        d[nc] = d[nc].astype(str).map(lambda x: mmap.get(x, x))
        dim_merged_any = True
    if dim_merged_any:
        infos.append("Учтено укрупнение категорий по переменным Росстата (таблица).")

    joint = rstat_f.groupby(weight_vars, dropna=False)[tgt_col].sum().reset_index()
    joint["count"] = joint[tgt_col].astype(float)
    total_j = float(joint["count"].sum())
    if total_j <= 0:
        return None, None, "Сумма целевых значений в файле Росстата должна быть больше нуля.", [], []

    tuple_keys = []
    for _, row in joint.iterrows():
        t = tuple(str(row[v]) for v in weight_vars)
        tuple_keys.append(t)
    joint["_tuple"] = tuple_keys
    pt_map: dict[tuple[str, ...], float] = {}
    for _, row in joint.iterrows():
        key = tuple(str(row[v]) for v in weight_vars)
        pt_map[key] = float(row["count"]) / total_j

    s_pt = sum(pt_map.values())
    if abs(s_pt - 1.0) > 1e-6:
        infos.append("Целевые доли Росстата не суммируются в 1, выполнена автоматическая нормировка.")
        pt_map = {k: float(v) / s_pt for k, v in pt_map.items()}

    jkeys = set(pt_map.keys())
    d["_sk"] = apply_merge_edges(build_stratum_key(d, [survey_col_map[v] for v in weight_vars]), merge_edges)

    stratum_per_row = d["_sk"].astype(str).tolist()
    row_cell_keys = [tuple(d.iloc[i][nc] for nc in norm_cols) for i in range(n)]

    for _, row in d.drop_duplicates(subset=norm_cols).iterrows():
        key = tuple(str(row[nc]) for nc in norm_cols)
        if key not in jkeys and any(key):
            warns.append("Не найдено соответствие для части категорий в данных и целевом файле.")

    w, preview_rows, warns2, infos2 = direct_weights_mode1_poststrat(n, stratum_per_row, row_cell_keys, pt_map)
    warns.extend(warns2)
    infos.extend(infos2)

    if np.any(~np.isfinite(w)):
        return (
            None,
            None,
            "Для части респондентов нельзя рассчитать вес. Проверьте соответствие категорий или объедините страты.",
            warns,
            infos,
        )

    preview = pd.DataFrame(preview_rows)
    return w, preview, "", warns, infos


def mode2_compute(
    survey: pd.DataFrame,
    split_var: str,
    target_cat: str,
    weighted_cat: str,
    soc_dem: list[str],
    merge_edges: list | None,
    na_is_category: bool,
) -> tuple[Optional[np.ndarray], Optional[pd.DataFrame], str]:
    if not soc_dem:
        return None, None, "Выберите хотя бы одну соц-дем переменную."
    if split_var not in survey.columns:
        return None, None, "В данных нет выбранной переменной-разделителя."

    d = survey.copy()
    for c in soc_dem:
        if c not in d.columns:
            return None, None, f"В данных нет колонки «{c}»."

    def split_val(x: Any) -> str:
        if is_no_text_cell(x):
            return "— Пропуск —" if na_is_category else ""
        if isinstance(x, str):
            return x.strip()
        return str(x).strip()

    d["_split_v"] = d[split_var].map(split_val)
    tgt_l = "— Пропуск —" if target_cat == "— Пропуск —" else str(target_cat).strip()
    wgt_l = "— Пропуск —" if weighted_cat == "— Пропуск —" else str(weighted_cat).strip()

    is_t = d["_split_v"] == tgt_l
    is_w = d["_split_v"] == wgt_l

    sk = build_stratum_key(d, soc_dem)
    sk = apply_merge_edges(sk, merge_edges)
    n = len(d)
    n_t = int(is_t.sum())
    n_w = int(is_w.sum())
    if n_t == 0 or n_w == 0:
        return None, None, "В выборке нет респондентов целевой или взвешиваемой группы."

    idx_by_s: dict[str, pd.Index] = {}
    for s in sk.unique():
        idx_by_s[str(s)] = sk[sk == s].index

    raw_by_stratum: dict[str, float] = {}
    for s, idx in idx_by_s.items():
        if stratum_is_proskip_only(s):
            raw_by_stratum[s] = 1.0
            continue
        nt = int(is_t.loc[idx].sum())
        nw = int(is_w.loc[idx].sum())
        if nw == 0:
            # В страте нет взвешиваемых (например только «пропуск» по разделителю — не выполнили условие).
            # Корректировать некого; коэффициент на респондентов W здесь не вешается — заглушка для словаря.
            raw_by_stratum[s] = 1.0
            continue
        pt = (nt / n_t) if n_t > 0 else 0.0
        pw = nw / n_w
        raw_by_stratum[s] = pt / pw if pw > 0 else np.nan

    w = np.ones(n, dtype=float)
    for i in range(n):
        s = str(sk.iloc[i])
        if is_w.iloc[i]:
            w[i] = raw_by_stratum.get(s, 1.0)

    if np.any(~np.isfinite(w)):
        return None, None, "Не удалось согласовать веса для всех страт. Попробуйте объединить страты."

    m = float(np.mean(w))
    if m <= 0 or not np.isfinite(m):
        w = np.ones(n, dtype=float)
    else:
        w = w / m
    for i in range(n):
        if is_t.iloc[i]:
            w[i] = 1.0
        if stratum_is_proskip_only(str(sk.iloc[i])):
            w[i] = 1.0

    rows = []
    strata = sorted(sk.unique(), key=str)
    for s in strata:
        if stratum_is_proskip_only(s):
            continue
        idx = sk[sk == s].index
        nt = int(is_t.loc[idx].sum())
        nw = int(is_w.loc[idx].sum())
        pt = nt / n_t if n_t else 0.0
        pw = nw / n_w if n_w else 0.0
        if nw == 0:
            rw = float("nan")
        else:
            rw = raw_by_stratum.get(str(s), np.nan)
        rows.append(
            {
                "stratum": s,
                "n_target": nt,
                "n_weighted": nw,
                "share_target": pt,
                "share_weighted": pw,
                "raw_w": rw,
            }
        )
    preview = pd.DataFrame(rows)
    return w, preview, ""


def natural_sort_key(s: str) -> tuple[Any, ...]:
    parts = re.split(r"(\d+)", str(s).lower())
    out: list[Any] = []
    for p in parts:
        if p.isdigit():
            out.append(int(p))
        else:
            out.append(p)
    return tuple(out)


def merge_edges_set_to_list(edge_set: set[frozenset[str]]) -> list[list[str]]:
    return [sorted(list(p), key=natural_sort_key) for p in edge_set]


def atoms_from_display(label: str) -> list[str]:
    return [p.strip() for p in str(label).split(" + ") if p.strip()]


def partner_atoms_for_display(a_display: str, edge_set: set[frozenset[str]]) -> set[str]:
    pa = set(atoms_from_display(a_display))
    out: set[str] = set()
    for e in edge_set:
        el = list(e)
        if len(el) != 2:
            continue
        u, v = el[0], el[1]
        if u in pa:
            out.add(v)
        if v in pa:
            out.add(u)
    return out


def display_for_atom(atom: str, all_strata: list[str]) -> str:
    for s in all_strata:
        if atom in atoms_from_display(s):
            return s
    return ""


def recommend_merge_partner(extreme: str, others: list[str]) -> str:
    if not others:
        return ""
    o_sorted = sorted(others, key=lambda x: natural_sort_key(x))
    if extreme in o_sorted:
        i = o_sorted.index(extreme)
        if i > 0:
            return o_sorted[i - 1]
        if i + 1 < len(o_sorted):
            return o_sorted[i + 1]
    best = o_sorted[0]
    ex_l = str(extreme).lower()
    best_score = -1
    for o in o_sorted:
        ol = o.lower()
        pref = 0
        for a, b in zip(ex_l, ol):
            if a == b:
                pref += 1
            else:
                break
        if pref > best_score:
            best_score = pref
            best = o
    return best


def ordered_merge_options(extreme: str, all_strata: list[str]) -> list[str]:
    others = [str(x) for x in all_strata if str(x) != str(extreme)]
    if not others:
        return []
    rec = recommend_merge_partner(extreme, others)
    rest = sorted([x for x in others if x != rec], key=natural_sort_key)
    if rec and rec in others:
        return [rec] + rest
    return sorted(others, key=natural_sort_key)


def style_preview_mode1(df: pd.DataFrame) -> Any:
    disp = df.rename(
        columns={
            "stratum": "Ячейка",
            "n": "Респондентов в выборке",
            "share_act": "Фактическая доля",
            "share_tgt": "Целевая доля",
            "raw_w": "Вес",
        }
    )
    def _hl(s: pd.Series) -> list[str]:
        out = []
        for v in s:
            try:
                x = float(v)
            except (TypeError, ValueError):
                out.append("")
                continue
            if x < RAW_LOW or x > RAW_HIGH:
                out.append(EXTREME_BG)
            else:
                out.append("")
        return out

    return disp.style.format(
        {"Фактическая доля": "{:.4f}", "Целевая доля": "{:.4f}", "Вес": "{:.4f}"}
    ).apply(_hl, subset=["Вес"])


def style_preview_mode2(df: pd.DataFrame) -> Any:
    disp = df.rename(
        columns={
            "stratum": "Ячейка",
            "n_target": "Респондентов (целевая)",
            "n_weighted": "Респондентов (взвешиваемая)",
            "share_target": "Доля в целевой группе",
            "share_weighted": "Доля во взвешиваемой группе",
            "raw_w": "Вес",
        }
    )

    def _hl(s: pd.Series) -> list[str]:
        out = []
        for v in s:
            try:
                x = float(v)
            except (TypeError, ValueError):
                out.append("")
                continue
            if x < RAW_LOW or x > RAW_HIGH:
                out.append(EXTREME_BG)
            else:
                out.append("")
        return out

    return disp.style.format(
        {
            "Доля в целевой группе": "{:.4f}",
            "Доля во взвешиваемой группе": "{:.4f}",
            "Вес": "{:.4f}",
        }
    ).apply(_hl, subset=["Вес"])


def stable_key(s: str) -> str:
    return hashlib.md5(s.encode("utf-8")).hexdigest()[:12]


def render_mode1_dim_merge_matrix(ros: pd.DataFrame, exclude_rows: list[str], exclude_cols: list[str]) -> None:
    """Укрупнение категорий по строкам/столбцам матрицы Росстата (режим 1)."""
    mg = st.session_state.setdefault("mode1_dim_merge_groups", {})
    mg.setdefault("matrix_col", [])
    mg.setdefault("matrix_row", [])

    st.subheader("Укрупнение категорий Росстата")
    st.caption(
        "Объединённые категории суммируются в **одну целевую ячейку** (как «Москва + Санкт-Петербург»). "
        "Вес и **предпросмотр** считаются по ячейке целиком: одинаковый вес у всех респондентов из группы, "
        "фактическая доля — по их суммарному числу в выборке. Исключённые из расчёта строки/столбцы здесь не показываются."
    )

    cols_geo = [
        str(c)
        for c in ros.columns[1:]
        if str(c).strip() not in exclude_cols and str(c).strip().lower() != "общий итог"
    ]
    col_opts = sorted(cols_geo, key=natural_sort_key)

    st.markdown("**Столбцы матрицы (география, тип НП и т.п.)**")
    pick_c = st.multiselect(
        "Выберите категории для одной группы",
        options=col_opts,
        key="m1_dim_col_pick",
    )
    if st.button("Добавить объединение по столбцам", key="m1_dim_col_add"):
        if len(pick_c) >= 2:
            mg["matrix_col"].append(list(pick_c))
            st.rerun()
        else:
            st.warning("Отметьте не менее двух столбцов.")

    for i, g in enumerate(list(mg["matrix_col"])):
        c1, c2 = st.columns([5, 1])
        with c1:
            st.text(" + ".join(g))
        with c2:
            if st.button("Удалить", key=f"m1_dim_col_rm_{i}"):
                mg["matrix_col"].pop(i)
                st.rerun()

    with st.expander("Укрупнение строк матрицы (при необходимости)", expanded=False):
        r0 = ros.iloc[:, 0].dropna().astype(str).unique().tolist()
        row_opts = sorted([x for x in r0 if x.strip() not in exclude_rows], key=natural_sort_key)
        pick_r = st.multiselect(
            "Строки в одну группу",
            options=row_opts,
            key="m1_dim_row_pick",
        )
        if st.button("Добавить объединение по строкам", key="m1_dim_row_add"):
            if len(pick_r) >= 2:
                mg["matrix_row"].append(list(pick_r))
                st.rerun()
            else:
                st.warning("Отметьте не менее двух строк.")
        for i, g in enumerate(list(mg["matrix_row"])):
            c1, c2 = st.columns([5, 1])
            with c1:
                st.text(" + ".join(g))
            with c2:
                if st.button("Удалить", key=f"m1_dim_row_rm_{i}"):
                    mg["matrix_row"].pop(i)
                    st.rerun()

    if st.button("Сбросить все укрупнения категорий", key="m1_dim_reset"):
        mg["matrix_col"] = []
        mg["matrix_row"] = []
        st.rerun()


def render_mode1_dim_merge_table(
    ros: pd.DataFrame,
    weight_vars: list[str],
    survey_col_map: dict[str, str],
    survey: pd.DataFrame,
    exclude_values: dict[str, list[str]],
) -> None:
    """Укрупнение по одной из переменных длинной таблицы Росстата."""
    if not weight_vars:
        return
    mg = st.session_state.setdefault("mode1_dim_merge_groups", {})
    for v in weight_vars:
        mg.setdefault(v, [])

    st.subheader("Укрупнение категорий Росстата")
    st.caption(
        "Выберите измерение и несколько категорий, которые нужно слить в одну целевую группу "
        "(численности в Росстате суммируются). Уже вошедшие в группы категории в списке не показываются. "
        "**Предпросмотр** — по укрупнённым ячейкам; вес одинаковый внутри ячейки."
    )

    dim = st.selectbox(
        "Измерение для укрупнения",
        options=list(weight_vars),
        format_func=lambda v: f"{v} → {survey_col_map.get(v, '')}",
        key="m1_dim_table_dim",
    )

    excl = {normalize_key(x) for x in (exclude_values.get(dim, []) or [])}
    merged_taken = _norm_keys_for_merged_groups(dim, mg.get(dim, []))

    rlist = _unique_ui_categories(ros[dim])
    scol = survey_col_map.get(dim)
    slist = _unique_ui_categories(survey[scol]) if scol in survey.columns else []
    opt_set: set[str] = set()
    for o in rlist + slist:
        nk = normalize_key(o)
        if nk in excl:
            continue
        if nk in merged_taken:
            continue
        opt_set.add(o)
    options = sorted(opt_set, key=natural_sort_key)

    pick = st.multiselect(
        f"Категории для объединения ({dim})",
        options=options,
        key=f"m1_dim_table_pick_{stable_key(str(dim))}",
    )
    if st.button("Добавить объединение", key="m1_dim_table_add"):
        if len(pick) >= 2:
            mg[dim].append(list(pick))
            st.rerun()
        else:
            st.warning("Отметьте не менее двух категорий.")

    for i, g in enumerate(list(mg.get(dim, []))):
        c1, c2 = st.columns([5, 1])
        with c1:
            st.text(" + ".join(g))
        with c2:
            if st.button("Удалить", key=f"m1_dim_tbl_rm_{stable_key(str(dim))}_{i}"):
                mg[dim].pop(i)
                st.rerun()

    if st.button("Сбросить укрупнения по всем измерениям", key="m1_dim_table_reset_all"):
        for v in weight_vars:
            mg[v] = []
        st.rerun()


def render_merge_controls(preview: pd.DataFrame, raw_col: str, key_prefix: str) -> list[list[str]]:
    extreme = preview[(preview[raw_col].notna()) & ((preview[raw_col] < RAW_LOW) | (preview[raw_col] > RAW_HIGH))]
    edge_set = merge_edges_list_to_set(st.session_state.merge_edges)
    if extreme.empty:
        return merge_edges_set_to_list(edge_set)
    st.markdown("**Объединение страт** (только в этой сессии; исходный файл не меняется)")
    st.caption(
        "Для каждой ячейки с экстремальным весом выберите вторую страту или пункт «Не объединять». "
        "После «Применить объединение» в предпросмотре появляется **одна новая группа** (название вида «A + B»), "
        "суммарная целевая доля и вес считаются по объединённым стратам."
    )
    all_strata = [str(x) for x in preview["stratum"].values]
    for _, r in extreme.iterrows():
        a = str(r["stratum"])
        base_opts = ordered_merge_options(a, all_strata)
        if not base_opts:
            continue
        rec = base_opts[0]
        options = [NO_MERGE_STRATUM] + base_opts
        partners = partner_atoms_for_display(a, edge_set)
        default_choice = NO_MERGE_STRATUM
        for atom in sorted(partners, key=natural_sort_key):
            disp = display_for_atom(atom, all_strata)
            if disp and disp != a:
                default_choice = disp
                break
        st.caption(f"Рекомендуется объединить «{a}» с «{rec}» (близкая по смыслу категория в отсортированном списке).")
        idx0 = options.index(default_choice) if default_choice in options else 0
        choice = st.selectbox(
            f"Объединить «{a}» с…",
            options=options,
            index=idx0,
            key=f"{key_prefix}_merge_{stable_key(a)}",
        )
        pa = set(atoms_from_display(a))
        edge_set = {e for e in edge_set if not (e & pa)}
        if choice != NO_MERGE_STRATUM:
            pb = set(atoms_from_display(choice))
            ra = sorted(pa, key=natural_sort_key)[0]
            rb = sorted(pb, key=natural_sort_key)[0]
            if ra != rb:
                edge_set.add(frozenset({ra, rb}))
    return merge_edges_set_to_list(edge_set)


def build_excel_with_wt(
    survey_bytes: bytes,
    header_row: int,
    weights: np.ndarray,
    leading_empty_rows: int = 0,
) -> bytes:
    """Добавляет столбец весов в конец листа: при двухстрочном заголовке — «WT» над строкой имён, «веса» в строке заголовков.

    leading_empty_rows — сколько полностью пустых строк шло сразу под заголовком в исходном файле;
    их мы не загружаем в DataFrame, поэтому первый вес сдвигается на столько строк вниз.
    """
    bio = io.BytesIO(survey_bytes)
    wb = load_workbook(bio)
    ws = wb.active
    if ws is None:
        raise ValueError("Пустая книга Excel.")

    hr = int(header_row)
    hdr = hr + 1
    max_c = ws.max_column or 1

    wt_col = max_c + 1
    if hr >= 1:
        ws.cell(row=hdr - 1, column=wt_col, value="WT")
        ws.cell(row=hdr, column=wt_col, value="веса")
    else:
        ws.cell(row=hdr, column=wt_col, value="веса")

    w = np.asarray(weights, dtype=float).ravel()
    skip = max(0, int(leading_empty_rows))
    data_start = hdr + 1 + skip
    for i in range(len(w)):
        ws.cell(row=data_start + i, column=wt_col, value=float(w[i]))

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def export_fallback_dataframe(survey_df: pd.DataFrame, weights: np.ndarray) -> bytes:
    out = survey_df.copy()
    n = len(out.columns)
    out.columns = pd.MultiIndex.from_arrays([list(out.columns), [""] * n])
    out[("WT", "веса")] = weights
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        out.to_excel(writer, index=False)
    return buf.getvalue()


def main() -> None:
    st.set_page_config(
        page_title="Взвешивание опросов",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    _init_session()

    with st.sidebar:
        st.title("Взвешивание опросов")
        data_file = st.file_uploader("Excel с данными опроса", type=["xlsx"], key="survey_up")

        hdr = st.number_input(
            "Строка с названиями колонок (переменные)",
            min_value=1,
            max_value=50,
            value=int(st.session_state.survey_header_row) + 1,
            step=1,
            help="Номер строки в Excel, по которой pandas называет колонки (часто 2). Строка **над ней** (например, 1) может содержать коды **Qgender**, **Qagerange**, **Qrucitysize** для табличного режима Росстата — сопоставление идёт по ним, имена колонок могут быть любыми.",
        )
        header_row = int(hdr) - 1

        mode = st.radio(
            "Режим",
            ("Внешние цели (Росстат)", "Группа на группу"),
        )

        if mode == "Внешние цели (Росстат)":
            if st.session_state.rosstat_df is None:
                rf = st.file_uploader("Файл Росстата (целевое распределение)", type=["xlsx"], key="ros_up")
                if rf is not None:
                    try:
                        st.session_state.rosstat_df = read_rosstat_excel(rf)
                        st.session_state.rosstat_name = rf.name
                    except Exception:
                        st.error("Не удалось прочитать файл Росстата. Проверьте формат.")
            else:
                st.success(f"Росстат загружен: **{st.session_state.rosstat_name}**")
                if st.button("Сменить файл Росстата"):
                    st.session_state.rosstat_df = None
                    st.session_state.rosstat_name = None
                    st.rerun()

        if st.button("Сбросить всё"):
            reset_session()
            st.rerun()

        with st.expander("Справка: как читать веса"):
            st.markdown(
                """
**Режим Росстат** — постстратификация: в ячейке вес = целевая доля ÷ фактическая доля в выборке.
Взвешенное распределение по ячейкам совпадает с целевым (с учётом исключений). Сумма весов обычно равна **N**,
если все респонденты в ячейках с целью и сумма целевых долей по ним равна 1. Нормировка «средний вес = 1» **не применяется**.

**Режим группа на группу** — эталон — одна подгруппа опроса; взвешивается другая так, чтобы по выбранным соц-дем
переменным её структура совпала с эталоном. У **целевой группы (эталона) вес всегда 1**; у взвешиваемой — поправки по стратам с тем же масштабом, что и раньше, но без сдвига веса эталона.
"""
            )

    if data_file is None and st.session_state.survey_bytes is None:
        st.info("Загрузите в боковой панели Excel с данными опроса.")
        return

    if data_file is not None:
        try:
            st.session_state.survey_bytes = data_file.getvalue()
            st.session_state.survey_name = getattr(data_file, "name", "опрос.xlsx")
        except Exception:
            data_file.seek(0)
            st.session_state.survey_bytes = data_file.read()
            st.session_state.survey_name = getattr(data_file, "name", "опрос.xlsx")

    st.session_state.survey_header_row = header_row

    try:
        survey = load_survey_excel(io.BytesIO(st.session_state.survey_bytes), header_row)
    except Exception:
        st.error("Не удалось прочитать файл с опросом. Проверьте, что это корректный Excel.")
        return

    st.session_state.survey_df = survey
    st.session_state.survey_leading_empty_rows = int(survey.attrs.get("leading_fully_empty_rows", 0))
    survey = st.session_state.survey_df

    if st.session_state.get("active_mode") != mode:
        st.session_state.active_mode = mode
        st.session_state.weights = None
        st.session_state.preview_df = None
        st.session_state.merge_edges = []
        st.session_state.mode1_dim_merge_groups = {}
        st.session_state.mode1_recalc = None
        st.session_state.mode2_params = None
        st.session_state.preview_mode = None
        st.session_state.preview_baseline = None

    if mode == "Внешние цели (Росстат)":
        st.header("Режим 1: подгонка к целям Росстата")
        ros = st.session_state.get("rosstat_df")
        if ros is None:
            st.warning("Загрузите файл Росстата в боковой панели.")
            return

        is_matrix = detect_rosstat_matrix(ros)
        st.caption(
            "Формат **матрицы**: в строках — пол и возраст (или склеенная подпись), в столбцах — география; в ячейках — численность."
            if is_matrix
            else "Формат **таблицы**: столбцы — измерения и показатель (численность или доля)."
        )

        exclude_rows: list[str] = []
        exclude_cols: list[str] = []
        exclude_long: dict[str, list[str]] = {}

        if is_matrix:
            r0 = ros.iloc[:, 0].dropna().astype(str).unique().tolist()
            cols = [str(c) for c in ros.columns[1:]]
            st.subheader("Группы Росстата, не участвующие в расчёте")
            st.caption("Отметьте строки и столбцы матрицы, которые нужно исключить из целевых долей.")
            exclude_rows = st.multiselect("Исключить строки матрицы", options=sorted(r0, key=natural_sort_key), default=[])
            exclude_cols = st.multiselect("Исключить столбцы (география и т.п.)", options=sorted(cols, key=natural_sort_key), default=[])

            st.subheader("Сопоставление с опросом")
            st.caption("Названия колонок опроса берутся из выбранной строки заголовков (см. боковую панель).")
            row_cols = st.multiselect(
                "Колонки опроса для **строк** матрицы (пол, возраст — можно несколько, значения склеиваются)",
                options=list(survey.columns),
                default=[survey.columns[0]] if len(survey.columns) else [],
            )
            col_col = st.selectbox(
                "Колонка опроса для **столбцов** матрицы (география)",
                options=list(survey.columns),
            )

            render_mode1_dim_merge_matrix(ros, exclude_rows, exclude_cols)

            if st.button("Рассчитать веса (предпросмотр)", key="calc1m"):
                if len(row_cols) < 1:
                    st.error("Выберите хотя бы одну колонку, соответствующую строкам матрицы Росстата.")
                else:
                    w, prev, err, warns, infos = mode1_matrix_compute(
                        survey,
                        ros,
                        row_cols,
                        col_col,
                        st.session_state.merge_edges,
                        exclude_rows,
                        exclude_cols,
                        copy.deepcopy(st.session_state.get("mode1_dim_merge_groups", {})),
                    )
                    for i in infos:
                        st.info(i)
                    for wn in warns:
                        st.warning(wn)
                    if err:
                        st.error(err)
                    else:
                        st.session_state.merge_edges = []
                        st.session_state.weights = w
                        st.session_state.preview_df = prev
                        st.session_state.preview_baseline = prev.copy()
                        st.session_state.preview_mode = 1
                        st.session_state.mode1_recalc = {
                            "kind": "matrix",
                            "row_cols": list(row_cols),
                            "col_col": col_col,
                            "exclude_rows": list(exclude_rows),
                            "exclude_cols": list(exclude_cols),
                            "dim_merge_groups": copy.deepcopy(
                                st.session_state.get("mode1_dim_merge_groups", {})
                            ),
                        }
                        st.session_state.last_warnings = warns
                        st.session_state.last_info = infos
                        st.rerun()

        else:
            std_wv, std_map, std_err = resolve_mode1_standard_table(ros, survey)
            if std_err:
                st.error(std_err)
                st.caption(
                    "В Росстате — колонки **Пол**, **Возраст**, **Тип населённого пункта** "
                    "(допускаются близкие названия). В опросе в **первой строке файла** (над строкой с названиями колонок) "
                    "должны быть ячейки **Qgender**, **Qagerange**, **Qrucitysize** в тех же столбцах; "
                    "если их нет — допускается совпадение с именем колонки во второй строке."
                )
            else:
                weight_vars = std_wv
                survey_col_map = std_map

                st.subheader("Сопоставление с опросом")
                st.info(
                    "Поля **Qgender** / **Qagerange** / **Qrucitysize** ищутся в **первой строке** листа опроса "
                    "(над строкой с названиями колонок); данные читаются по колонкам со второй строки, как раньше. "
                    "Столбцы Росстата подставляются автоматически."
                )
                pairs = " · ".join(f"«{r}» → {survey_col_map[r]}" for r in weight_vars)
                st.success(f"Найдено сопоставление: {pairs}")

                exclude_long = {}
                st.subheader("Группы Росстата, не участвующие в расчёте")
                for v in weight_vars:
                    vals = _unique_ui_categories(ros[v])
                    if vals:
                        exclude_long[v] = st.multiselect(
                            f"Исключить значения в «{v}»",
                            options=vals,
                            key=f"ex_{stable_key(str(v))}",
                        )

                render_mode1_dim_merge_table(ros, weight_vars, survey_col_map, survey, exclude_long)

                if st.button("Рассчитать веса (предпросмотр)", key="calc1t"):
                    try:
                        w, prev, err, warns, infos = mode1_table_compute(
                            survey,
                            ros,
                            weight_vars,
                            survey_col_map,
                            st.session_state.merge_edges,
                            exclude_long,
                            copy.deepcopy(st.session_state.get("mode1_dim_merge_groups", {})),
                        )
                    except Exception:
                        st.error(
                            "Не удалось выполнить расчёт. Проверьте, что в файле Росстата есть численности или доли "
                            "и что категории в опросе совпадают по смыслу с категориями в файле Росстата."
                        )
                    else:
                        for i in infos:
                            st.info(i)
                        for wn in warns:
                            st.warning(wn)
                        if err:
                            st.error(err)
                        else:
                            st.session_state.merge_edges = []
                            st.session_state.weights = w
                            st.session_state.preview_df = prev
                            st.session_state.preview_baseline = prev.copy()
                            st.session_state.preview_mode = 1
                            st.session_state.mode1_recalc = {
                                "kind": "table",
                                "weight_vars": list(weight_vars),
                                "survey_col_map": dict(survey_col_map),
                                "exclude_long": {k: list(v) for k, v in exclude_long.items()},
                                "dim_merge_groups": copy.deepcopy(
                                    st.session_state.get("mode1_dim_merge_groups", {})
                                ),
                            }
                            st.session_state.last_warnings = warns
                            st.session_state.last_info = infos
                            st.rerun()

        prev1 = st.session_state.preview_df
        if prev1 is not None and st.session_state.get("preview_mode") == 1:
            merge_on = bool(st.session_state.get("merge_edges"))
            base = st.session_state.get("preview_baseline")
            if merge_on and base is not None and len(base) > 0:
                with st.expander("Предпросмотр до объединения страт", expanded=False):
                    st.caption("Исходные ячейки; экстремальные веса (ниже 0,3 или выше 2,5) подсвечены.")
                    st.dataframe(style_preview_mode1(base), use_container_width=True, hide_index=True)
            st.subheader(
                "Предпросмотр после объединения страт" if merge_on else "Предпросмотр по ячейкам"
            )
            if merge_on:
                st.caption(
                    "Таблица **пересчитана** по укрупнённым ячейкам; веса и доли соответствуют выбранному объединению страт."
                )
            else:
                st.caption("Экстремальные веса (ниже 0,3 или выше 2,5) подсвечены красным. При необходимости объедините страты ниже.")
            if st.session_state.weights is not None:
                ww = np.asarray(st.session_state.weights, dtype=float)
                st.caption(
                    f"Контроль подгонки: сумма весов = **{np.nansum(ww):.2f}**, число респондентов **{len(ww)}**. "
                    "При полном покрытии ячеек и сумме целевых долей 1 они совпадают."
                )
            st.dataframe(style_preview_mode1(prev1), use_container_width=True, hide_index=True)

            new_edges = render_merge_controls(prev1, "raw_w", "m1")
            if st.button("Применить объединение", key="apply_merge1"):
                st.session_state.merge_edges = new_edges
                p = st.session_state.mode1_recalc
                sur = st.session_state.survey_df
                ros2 = st.session_state.rosstat_df
                if p and sur is not None and ros2 is not None:
                    if p["kind"] == "matrix":
                        w, prev2, err, warns, infos = mode1_matrix_compute(
                            sur,
                            ros2,
                            p["row_cols"],
                            p["col_col"],
                            new_edges,
                            p.get("exclude_rows", []),
                            p.get("exclude_cols", []),
                            p.get("dim_merge_groups") or {},
                        )
                    else:
                        w, prev2, err, warns, infos = mode1_table_compute(
                            sur,
                            ros2,
                            p["weight_vars"],
                            p["survey_col_map"],
                            new_edges,
                            p.get("exclude_long", {}),
                            p.get("dim_merge_groups") or {},
                        )
                    for i in infos:
                        st.info(i)
                    for wn in warns:
                        st.warning(wn)
                    if err:
                        st.error(err)
                    else:
                        st.session_state.weights = w
                        st.session_state.preview_df = prev2
                st.rerun()

            if st.session_state.weights is not None:
                wb = st.session_state.weights
                try:
                    xbytes = build_excel_with_wt(
                        st.session_state.survey_bytes,
                        header_row,
                        wb,
                        st.session_state.get("survey_leading_empty_rows", 0),
                    )
                    fname = (st.session_state.survey_name or "опрос").rsplit(".", 1)[0] + "_с_WT.xlsx"
                    st.download_button(
                        label="Применить веса и скачать Excel",
                        data=xbytes,
                        file_name=fname,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl1",
                    )
                except Exception:
                    st.caption("Не удалось сохранить исходный макет листа; выгружается таблица с заголовками из строки переменных.")
                    xbytes = export_fallback_dataframe(survey, wb)
                    st.download_button(
                        label="Применить веса и скачать Excel",
                        data=xbytes,
                        file_name="опрос_с_весами_WT.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl1b",
                    )

    else:
        st.header("Режим 2: группа на группу")
        split_var = st.selectbox("Переменная-разделитель", options=list(survey.columns))
        vals = survey[split_var].unique()
        cat_list: list[str] = []

        def _ui_cat(v: Any) -> str:
            if is_no_text_cell(v):
                return "— Пропуск —"
            if isinstance(v, str):
                return v.strip()
            return str(v).strip()

        any_no_text = bool(survey[split_var].map(is_no_text_cell).any())
        for v in vals:
            cat_list.append(_ui_cat(v))
        if any_no_text and "— Пропуск —" not in cat_list:
            cat_list.append("— Пропуск —")
        cat_list = sorted(set(cat_list), key=lambda x: (x != "— Пропуск —", natural_sort_key(x)))

        target_cat = st.selectbox("Целевая группа (эталон)", options=cat_list)
        weighted_cat = st.selectbox("Взвешиваемая группа", options=cat_list)
        soc_dem = st.multiselect(
            "Соц-дем переменные (страты)",
            options=[c for c in survey.columns if c != split_var],
        )

        if st.button("Рассчитать веса (предпросмотр)", key="calc2"):
            if target_cat == weighted_cat:
                st.error("Целевая и взвешиваемая группы должны различаться.")
            else:
                w, prev, err = mode2_compute(
                    survey,
                    split_var,
                    target_cat,
                    weighted_cat,
                    soc_dem,
                    st.session_state.merge_edges,
                    na_is_category=True,
                )
                if err:
                    st.error(err)
                else:
                    st.session_state.merge_edges = []
                    st.session_state.weights = w
                    st.session_state.preview_df = prev
                    st.session_state.preview_baseline = prev.copy()
                    st.session_state.preview_mode = 2
                    st.session_state.mode2_params = {
                        "split_var": split_var,
                        "target_cat": target_cat,
                        "weighted_cat": weighted_cat,
                        "soc_dem": list(soc_dem),
                    }
                    st.rerun()

        prev2 = st.session_state.preview_df
        if prev2 is not None and st.session_state.get("preview_mode") == 2:
            merge_on = bool(st.session_state.get("merge_edges"))
            base2 = st.session_state.get("preview_baseline")
            if merge_on and base2 is not None and len(base2) > 0:
                with st.expander("Предпросмотр до объединения страт", expanded=False):
                    st.caption("Исходные ячейки; экстремальные веса подсвечены.")
                    st.dataframe(style_preview_mode2(base2), use_container_width=True, hide_index=True)
            st.subheader(
                "Предпросмотр после объединения страт" if merge_on else "Предпросмотр по ячейкам"
            )
            if merge_on:
                st.caption(
                    "Таблица **пересчитана** по укрупнённым ячейкам после объединения; веса обновлены."
                )
            else:
                st.caption("Экстремальные веса (ниже 0,3 или выше 2,5) подсвечены красным.")
            st.dataframe(style_preview_mode2(prev2), use_container_width=True, hide_index=True)

            new_edges = render_merge_controls(prev2, "raw_w", "m2")
            if st.button("Применить объединение", key="apply_merge2"):
                st.session_state.merge_edges = new_edges
                p = st.session_state.mode2_params
                sur = st.session_state.survey_df
                if p and sur is not None:
                    w, prev2b, err = mode2_compute(
                        sur,
                        p["split_var"],
                        p["target_cat"],
                        p["weighted_cat"],
                        p["soc_dem"],
                        new_edges,
                        na_is_category=True,
                    )
                    if err:
                        st.error(err)
                    else:
                        st.session_state.weights = w
                        st.session_state.preview_df = prev2b
                st.rerun()

            if st.session_state.weights is not None:
                wb = st.session_state.weights
                try:
                    xbytes = build_excel_with_wt(
                        st.session_state.survey_bytes,
                        header_row,
                        wb,
                        st.session_state.get("survey_leading_empty_rows", 0),
                    )
                    fname = (st.session_state.survey_name or "опрос").rsplit(".", 1)[0] + "_с_WT.xlsx"
                    st.download_button(
                        label="Применить веса и скачать Excel",
                        data=xbytes,
                        file_name=fname,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl2",
                    )
                except Exception:
                    st.caption("Не удалось сохранить исходный макет листа; выгружается таблица с заголовками из строки переменных.")
                    xbytes = export_fallback_dataframe(survey, wb)
                    st.download_button(
                        label="Применить веса и скачать Excel",
                        data=xbytes,
                        file_name="опрос_с_весами_WT.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl2b",
                    )


if __name__ == "__main__":
    main()
